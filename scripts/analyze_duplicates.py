#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUT = ROOT / "outputs"
OUT.mkdir(parents=True, exist_ok=True)


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000D_", "").strip()


def num(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(str(value).replace(",", "").strip())
    except ValueError:
        return None


def is_excluded_unit(emd: str, unit: str) -> bool:
    joined = f"{emd} {unit}"
    excluded = ["합계", "계", "소계", "거소", "선상", "국외", "재외"]
    return any(token in joined for token in excluded)


def vote_class(emd: str, unit: str) -> str | None:
    joined = f"{emd} {unit}"
    if "관내사전투표" in joined:
        return "관내사전투표"
    if "관외사전투표" in joined:
        return "관외사전투표"
    if "국내부재자투표" in joined:
        return "국내부재자투표"
    return None


def candidate_name(label: str) -> str:
    parts = [p.strip() for p in clean(label).splitlines() if p.strip()]
    return parts[-1] if parts else clean(label)


def parse_candidate_columns(headers: Iterable[Any], start: int) -> list[tuple[int, str]]:
    cols = []
    for idx, raw in enumerate(headers[start:], start):
        label = clean(raw)
        if not label or label == "/" or label in {"계", "무효", "기권", "무효 투표수", "기권자수"}:
            continue
        if label.startswith("후보") or label == "후보자별 득표수":
            continue
        cols.append((idx, label))
    return cols


def parse_assembly(path: Path, election: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["지역구"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    i = 1
    while i < len(rows):
        row = rows[i]
        if clean(row[2] if len(row) > 2 else "") == "" and num(row[6] if len(row) > 6 else None) is None:
            parties = [clean(x) for x in row[6:]]
            names = [clean(x) for x in rows[i + 1][6:]]
            labels = []
            for party, name in zip(parties, names):
                if party or name:
                    labels.append(f"{party}\n{name}".strip())
                else:
                    labels.append("")
            ccols = [(6 + j, label) for j, label in enumerate(labels) if label]
            i += 2
            last_emd: dict[tuple[str, str], str] = {}
            while i < len(rows):
                r = rows[i]
                if clean(r[0]) and clean(r[2]) == "" and clean(r[3]) == "" and num(r[6] if len(r) > 6 else None) is None:
                    break
                sido, district = clean(r[0]), clean(r[1])
                emd, unit = clean(r[2]), clean(r[3])
                if emd and unit == "소계":
                    last_emd[(sido, district)] = emd
                effective_emd = emd or last_emd.get((sido, district), "")
                vc = vote_class(effective_emd, unit)
                if vc and not is_excluded_unit(effective_emd, unit):
                    votes = [num(r[idx]) or 0 for idx, _ in ccols]
                    out.append({
                        "election": election,
                        "category": "국회의원",
                        "contest": f"{sido}|{district}",
                        "unit": f"{effective_emd}|{unit}",
                        "vote_class": vc,
                        "candidate_labels": [label for _, label in ccols],
                        "votes": votes,
                    })
                i += 1
            continue
        i += 1
    return out


def parse_presidential(path: Path, election: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if clean(r[0]) in {"시도", "시도명"})
    cand_start = 6
    if any("후보자별" in clean(x) for x in rows[header_idx]):
        cand_row = rows[header_idx + 1]
        data_start = header_idx + 2
    else:
        cand_row = rows[header_idx]
        data_start = header_idx + 1
    ccols = parse_candidate_columns(cand_row, cand_start)
    out = []
    last_emd: dict[tuple[str, str], str] = {}
    current_sido = ""
    current_district = ""
    for r in rows[data_start:]:
        if not any(clean(x) for x in r[:8]):
            continue
        raw_sido, raw_district = clean(r[0]), clean(r[1])
        if raw_sido:
            current_sido = raw_sido
        if raw_district:
            current_district = raw_district
        sido, district = current_sido, current_district
        emd, unit = clean(r[2]), clean(r[3])
        if not sido:
            continue
        if emd and unit in {"소계", "계"}:
            last_emd[(sido, district)] = emd
        effective_emd = emd or last_emd.get((sido, district), "")
        vc = vote_class(effective_emd, unit)
        if vc and not is_excluded_unit(effective_emd, unit):
            votes = [num(r[idx]) or 0 for idx, _ in ccols]
            out.append({
                "election": election,
                "category": "대통령",
                "contest": "전국",
                "unit": f"{sido}|{district}|{effective_emd}|{unit}",
                "vote_class": vc,
                "candidate_labels": [label for _, label in ccols],
                "votes": votes,
            })
    return out


def local_layout(row0: tuple[Any, ...], row2: tuple[Any, ...]) -> tuple[int, int, int, int, int]:
    labels = [clean(x) for x in row0]
    if labels[:4] == ["시도", "구시군", "읍면동", "구분"]:
        return 0, 0, 2, 3, 6
    if labels[:5] == ["시도", "선거구", "구시군", "읍면동", "구분"]:
        return 0, 1, 3, 4, 7
    if labels[:5] == ["시도", "구시군", "선거구", "읍면동", "구분"]:
        return 0, 2, 3, 4, 7
    if labels[:4] == ["선거구명", "구시군명", "읍면동명", "구분"]:
        return 0, 0, 2, 3, 6
    if labels[:5] == ["시도명", "구시군명", "선거구(구시군)", "읍면동명", "구분"]:
        return 0, 2, 3, 4, 7
    if labels[:6] == ["선거종류", "선거구명", "시도명", "구시군명", "읍면동명", "구분"]:
        return 2, 1, 4, 5, 8
    if labels[:7] == ["선거종류", "시도", "선거구명", "시도명", "구시군명", "읍면동명", "구분"]:
        return 1, 2, 5, 6, 9
    raise ValueError(f"unknown local layout: {labels[:10]}")


def parse_local(path: Path, election: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue
        try:
            sido_col, contest_col, emd_col, unit_col, cand_start = local_layout(rows[0], rows[2])
        except ValueError:
            continue
        ccols: list[tuple[int, str]] = []
        current_sido = ""
        current_contest = ""
        last_emd: dict[tuple[str, str], str] = {}
        for r in rows[1:]:
            candidate_like = any(clean(x) for x in r[cand_start:]) and all(num(x) is None for x in r[cand_start:] if clean(x))
            if candidate_like:
                parsed = parse_candidate_columns(r, cand_start)
                if parsed:
                    ccols = parsed
                    raw_sido = clean(r[sido_col])
                    raw_contest = clean(r[contest_col])
                    if raw_sido and raw_sido not in {"합계", "계"}:
                        current_sido = raw_sido
                    if raw_contest and raw_contest not in {"합계", "계"}:
                        current_contest = raw_contest
                continue
            if not ccols:
                continue
            raw_sido = clean(r[sido_col])
            raw_contest = clean(r[contest_col])
            if raw_sido and raw_sido not in {"합계", "계"}:
                current_sido = raw_sido
            if raw_contest and raw_contest not in {"합계", "계"}:
                current_contest = raw_contest
            sido = current_sido
            contest = current_contest or sido
            emd = clean(r[emd_col])
            unit = clean(r[unit_col])
            if not contest:
                continue
            if emd and unit in {"소계", "계"}:
                last_emd[(sido, contest)] = emd
            effective_emd = emd or last_emd.get((sido, contest), "")
            vc = vote_class(effective_emd, unit)
            if vc and not is_excluded_unit(effective_emd, unit):
                votes = [num(r[idx]) or 0 for idx, _ in ccols]
                out.append({
                    "election": election,
                    "category": sheet,
                    "contest": f"{sido}|{contest}",
                    "unit": f"{effective_emd}|{unit}",
                    "vote_class": vc,
                    "candidate_labels": [label for _, label in ccols],
                    "votes": votes,
                })
    return out


DATASETS = [
    ("2014-local", DATA / "local2014.xlsx", parse_local),
    ("2016-assembly", DATA / "assembly2016.xlsx", parse_assembly),
    ("2017-presidential", DATA / "pres2017.xlsx", parse_presidential),
    ("2018-local", DATA / "local2018.xlsx", parse_local),
    ("2020-assembly", DATA / "assembly2020.xlsx", parse_assembly),
    ("2022-presidential", DATA / "pres2022.xlsx", parse_presidential),
    ("2022-local", DATA / "local2022.xlsx", parse_local),
    ("2024-assembly", DATA / "assembly2024.xlsx", parse_assembly),
    ("2025-presidential", DATA / "pres2025.xlsx", parse_presidential),
]


def top2_signature(rec: dict[str, Any]) -> tuple[int, ...]:
    return tuple(rec["votes"][:2])


def all_signature(rec: dict[str, Any]) -> tuple[int, ...]:
    return tuple(rec["votes"])


def summarize(records: list[dict[str, Any]], sig_name: str, sig_fn) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    buckets = defaultdict(list)
    for rec in records:
        key = (rec["election"], rec["category"], rec["contest"], rec["vote_class"], sig_fn(rec))
        buckets[key].append(rec)
    groups = []
    examples = []
    for (election, category, contest, vc, sig), rows in buckets.items():
        if len(rows) < 2:
            continue
        groups.append({
            "signature_type": sig_name,
            "election": election,
            "category": category,
            "contest": contest,
            "vote_class": vc,
            "signature": "|".join(map(str, sig)),
            "duplicate_units": len(rows),
        })
        for rec in rows[:10]:
            examples.append({
                "signature_type": sig_name,
                "election": election,
                "category": category,
                "contest": contest,
                "vote_class": vc,
                "signature": "|".join(map(str, sig)),
                "unit": rec["unit"],
                "candidate_names": "|".join(candidate_name(x) for x in rec["candidate_labels"][:len(sig)]),
            })
    return groups, examples


def main() -> None:
    records = []
    dataset_counts = []
    for name, path, parser in DATASETS:
        parsed = parser(path, name)
        records.extend(parsed)
        dataset_counts.append({"dataset": name, "rows": len(parsed)})

    all_groups, all_examples = summarize(records, "all_candidates", all_signature)
    top2_groups, top2_examples = summarize(records, "top2_candidates", top2_signature)
    groups = all_groups + top2_groups
    examples = all_examples + top2_examples

    summary = defaultdict(lambda: {
        "units": 0,
        "all_candidate_duplicate_groups": 0,
        "all_candidate_duplicate_units": 0,
        "top2_duplicate_groups": 0,
        "top2_duplicate_units": 0,
    })
    for rec in records:
        summary[(rec["election"], rec["category"], rec["vote_class"])]["units"] += 1
    for g in all_groups:
        item = summary[(g["election"], g["category"], g["vote_class"])]
        item["all_candidate_duplicate_groups"] += 1
        item["all_candidate_duplicate_units"] += g["duplicate_units"]
    for g in top2_groups:
        item = summary[(g["election"], g["category"], g["vote_class"])]
        item["top2_duplicate_groups"] += 1
        item["top2_duplicate_units"] += g["duplicate_units"]

    summary_rows = [
        {"election": k[0], "category": k[1], "vote_class": k[2], **v}
        for k, v in sorted(summary.items())
    ]

    def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
        if not rows:
            path.write_text("", encoding="utf-8")
            return
        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)

    write_csv(OUT / "dataset_counts.csv", dataset_counts)
    write_csv(OUT / "duplicate_summary.csv", summary_rows)
    write_csv(OUT / "duplicate_groups.csv", sorted(groups, key=lambda r: (r["signature_type"], r["election"], -r["duplicate_units"])))
    write_csv(OUT / "duplicate_examples.csv", sorted(examples, key=lambda r: (r["signature_type"], r["election"], r["contest"], r["signature"], r["unit"])))
    (OUT / "manifest.json").write_text(json.dumps({"datasets": dataset_counts, "record_count": len(records)}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"datasets": dataset_counts, "record_count": len(records), "duplicate_groups": len(groups)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
